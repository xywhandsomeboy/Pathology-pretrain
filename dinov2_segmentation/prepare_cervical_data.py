"""Prepare a frozen, binary cervical WSI cohort from partially downloaded data.

The source annotations distinguish low grade, high grade and malignant lesions.
This pipeline intentionally maps every annotated lesion polygon to class 1 and
uses class 0 for background/normal tissue.  Unknown non-empty labels fail
closed so a new source label can never be silently assigned to the wrong class.
"""

from __future__ import annotations

import argparse
from concurrent.futures import ProcessPoolExecutor, as_completed
import csv
import hashlib
import json
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

import numpy as np
from PIL import Image, ImageDraw


TUMOR_CATEGORIES = {"low_grade", "high_grade", "malignant"}
NORMAL_CATEGORY = "normal_inflammation"
TUMOR_ANNOTATIONS = {"low grade", "high grade", "malignant"}
PATCH_FIELDS = (
    "slide_name_split",
    "slide_id",
    "patch_id",
    "filepath",
    "image_path",
    "mask_path",
    "x",
    "y",
    "level",
    "split",
    "has_tumor",
    "tumor_fraction",
    "tissue_fraction",
)


def _atomic_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def _atomic_csv(path: Path, rows: Iterable[dict], fieldnames=PATCH_FIELDS) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", newline="", encoding="utf-8") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(path)


def _expected_sizes(path: Path) -> dict[str, int]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise TypeError("Download manifest must be a JSON list")
    expected: dict[str, int] = {}
    for item in payload:
        manifest_name = Path(item["path"]).name
        # S-BIAD1168 has 2,348 entries such as ``slide-id .geojson`` while
        # the corresponding WSI and metadata use ``slide-id``. Normalize only
        # whitespace immediately before the final suffix so the downloaded
        # annotation remains pairable with ``slide-id.isyntax``.
        parsed_name = Path(manifest_name)
        name = f"{parsed_name.stem.rstrip()}{parsed_name.suffix}"
        size = int(item["size"])
        if name in expected and expected[name] != size:
            raise ValueError(
                f"Conflicting manifest sizes after filename normalization: {name}"
            )
        expected[name] = size
    return expected


def _is_exact_file(path: Path, expected: dict[str, int]) -> bool:
    expected_size = expected.get(path.name)
    # A zero-byte object can be a complete download when the upstream manifest
    # itself declares size 0, but it is not usable as a WSI, annotation or
    # preview image during data preparation.
    return (
        path.is_file()
        and expected_size is not None
        and expected_size > 0
        and path.stat().st_size == expected_size
    )


def _metadata_records(path: Path) -> dict[str, dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("openpyxl is required to read the official data split") from error
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    values = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(values)]
    required = {"Filename", "Category", "train/valid/test", "ExcludedFromAnnotation"}
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"Metadata workbook is missing columns: {missing}")
    records = {}
    for row_values in values:
        row = dict(zip(headers, row_values))
        if not row.get("Filename"):
            continue
        slide_id = Path(str(row["Filename"])).stem
        records[slide_id] = row
    return records


def _normalized_annotation(value: object) -> str:
    return " ".join(str(value).strip().lower().replace("_", " ").split())


def _annotation_features(path: Path) -> list[dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("type") != "FeatureCollection" or not isinstance(payload.get("features"), list):
        raise ValueError(f"Invalid GeoJSON FeatureCollection: {path}")
    features = payload["features"]
    for feature in features:
        name = (
            feature.get("properties", {})
            .get("classification", {})
            .get("name", "")
        )
        normalized = _normalized_annotation(name)
        if normalized not in TUMOR_ANNOTATIONS:
            raise ValueError(f"Unknown non-empty annotation label {name!r} in {path}")
    return features


def discover_ready(args) -> tuple[list[dict], dict]:
    expected = _expected_sizes(args.download_manifest)
    metadata = _metadata_records(args.metadata_xlsx)
    ready = []
    rejected = Counter()
    for slide_id, record in sorted(metadata.items()):
        category = str(record.get("Category") or "").strip().lower()
        split = str(record.get("train/valid/test") or "").strip().lower()
        excluded = str(record.get("ExcludedFromAnnotation") or "0").strip().lower()
        if category not in TUMOR_CATEGORIES | {NORMAL_CATEGORY}:
            rejected["unsupported_category"] += 1
            continue
        if split not in {"train", "valid", "test"}:
            rejected["unsupported_split"] += 1
            continue
        if excluded not in {"0", "0.0", "false", "none", ""}:
            rejected["excluded"] += 1
            continue
        slide_path = args.data_root / f"{slide_id}.isyntax"
        annotation_path = args.data_root / f"{slide_id}.geojson"
        if not _is_exact_file(slide_path, expected):
            rejected["incomplete_isyntax"] += 1
            continue
        if not _is_exact_file(annotation_path, expected):
            rejected["incomplete_geojson"] += 1
            continue
        features = _annotation_features(annotation_path)
        if category == NORMAL_CATEGORY and features:
            raise ValueError(f"Normal slide {slide_id} unexpectedly contains lesion polygons")
        if category in TUMOR_CATEGORIES and not features:
            raise ValueError(f"Tumor slide {slide_id} has no lesion polygons")
        ready.append(
            {
                "slide_id": slide_id,
                "split": split,
                "binary_slide_class": int(category in TUMOR_CATEGORIES),
                "isyntax_path": str(slide_path.resolve()),
                "geojson_path": str(annotation_path.resolve()),
                "isyntax_size": slide_path.stat().st_size,
                "geojson_size": annotation_path.stat().st_size,
            }
        )
    split_counts = Counter(record["split"] for record in ready)
    class_counts = Counter(record["binary_slide_class"] for record in ready)
    is_ready = (
        len(ready) >= args.minimum_ready
        and split_counts["train"] >= args.minimum_train
        and split_counts["valid"] >= args.minimum_valid
        and class_counts[0] >= 1
        and class_counts[1] >= 1
    )
    status = {
        "ready": is_ready,
        "ready_slides": len(ready),
        "minimum_ready": args.minimum_ready,
        "split_counts": dict(split_counts),
        "minimum_train": args.minimum_train,
        "minimum_valid": args.minimum_valid,
        "binary_slide_class_counts": {str(key): value for key, value in class_counts.items()},
        "rejected": dict(rejected),
        "label_map": {
            "normal_inflammation/background": 0,
            "low_grade": 1,
            "high_grade": 1,
            "malignant": 1,
        },
    }
    return ready, status


def _geometry_polygons(feature: dict) -> list[list[list[tuple[float, float]]]]:
    geometry = feature.get("geometry") or {}
    coordinates = geometry.get("coordinates")
    if geometry.get("type") == "Polygon":
        polygons = [coordinates]
    elif geometry.get("type") == "MultiPolygon":
        polygons = coordinates
    else:
        raise ValueError(f"Unsupported annotation geometry: {geometry.get('type')!r}")
    result = []
    for polygon in polygons or []:
        rings = []
        for ring in polygon or []:
            points = [(float(point[0]), float(point[1])) for point in ring]
            if len(points) >= 3:
                rings.append(points)
        if rings:
            result.append(rings)
    return result


def _scaled_polygons(features: list[dict], downsample: float) -> list[dict]:
    polygons = []
    for feature in features:
        for rings in _geometry_polygons(feature):
            scaled = [[(x / downsample, y / downsample) for x, y in ring] for ring in rings]
            all_points = [point for ring in scaled for point in ring]
            xs, ys = zip(*all_points)
            polygons.append({"rings": scaled, "bbox": (min(xs), min(ys), max(xs), max(ys))})
    return polygons


def _binary_mask(polygons: list[dict], x: int, y: int, size: int) -> Image.Image:
    mask = Image.new("L", (size, size), 0)
    draw = ImageDraw.Draw(mask)
    right, bottom = x + size, y + size
    for polygon in polygons:
        left, top, polygon_right, polygon_bottom = polygon["bbox"]
        if polygon_right < x or polygon_bottom < y or left >= right or top >= bottom:
            continue
        rings = polygon["rings"]
        draw.polygon([(px - x, py - y) for px, py in rings[0]], fill=1)
        for hole in rings[1:]:
            draw.polygon([(px - x, py - y) for px, py in hole], fill=0)
    return mask


def _preview_tissue(slide, preview_path: Path | None) -> np.ndarray:
    if preview_path is not None and preview_path.is_file():
        rgb = np.asarray(Image.open(preview_path).convert("RGB"))
        return np.any(rgb != 0, axis=2).astype(np.uint8)
    level = min(5, int(slide.level_count) - 1)
    width, height = map(int, slide.level_dimensions[level])
    rgb = slide.read_region(0, 0, width, height, level=level)[..., :3]
    maximum = rgb.max(axis=2)
    minimum = rgb.min(axis=2)
    return ((maximum < 245) & (minimum > 15) & ((maximum - minimum) > 5)).astype(np.uint8)


def _integral(array: np.ndarray) -> np.ndarray:
    return np.pad(array.astype(np.int64), ((1, 0), (1, 0))).cumsum(0).cumsum(1)


def _area_sum(integral: np.ndarray, left: int, top: int, right: int, bottom: int) -> int:
    return int(
        integral[bottom, right]
        - integral[top, right]
        - integral[bottom, left]
        + integral[top, left]
    )


def _grid_positions(length: int, patch_size: int, stride: int) -> list[int]:
    if length < patch_size:
        return []
    positions = list(range(0, length - patch_size + 1, stride))
    last = length - patch_size
    if positions[-1] != last:
        positions.append(last)
    return positions


def _valid_image(path: Path, size: int) -> bool:
    if not path.is_file():
        return False
    try:
        with Image.open(path) as image:
            return image.size == (size, size)
    except OSError:
        return False


def _save_patch(image: Image.Image, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    image.save(temporary, format="JPEG", quality=95, subsampling=0)
    temporary.replace(path)


def _save_mask(mask: Image.Image, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    mask.save(temporary, format="PNG", compress_level=3)
    temporary.replace(path)


def _process_slide(record: dict, args, expected: dict[str, int]) -> list[dict]:
    try:
        from isyntax import ISyntax
    except ImportError as error:
        raise RuntimeError("pyisyntax is required to read .isyntax slides") from error
    slide_id = record["slide_id"]
    slide_output = args.output_root / "slides" / slide_id
    rows_path = slide_output / "patches.csv"
    marker_path = slide_output / "complete.json"
    source_signature = {
        "slide_id": slide_id,
        "isyntax_size": record["isyntax_size"],
        "geojson_size": record["geojson_size"],
        "level": args.level,
        "patch_size": args.patch_size,
        "stride": args.stride,
        "minimum_tissue_fraction": args.minimum_tissue_fraction,
        "binary_label_version": 1,
    }
    if marker_path.is_file():
        existing = json.loads(marker_path.read_text(encoding="utf-8"))
        if existing.get("source_signature") != source_signature:
            raise ValueError(f"Existing prepared slide uses different settings: {slide_id}")
        with rows_path.open(newline="", encoding="utf-8") as stream:
            return list(csv.DictReader(stream))

    features = _annotation_features(Path(record["geojson_path"]))
    slide_path = Path(record["isyntax_path"])
    preview_path = args.data_root / f"{slide_id}_mask.png"
    if not _is_exact_file(preview_path, expected):
        preview_path = None
    with ISyntax.open(slide_path) as slide:
        if args.level < 0 or args.level >= int(slide.level_count):
            raise ValueError(f"Requested level {args.level} is absent from {slide_id}")
        level_width, level_height = map(int, slide.level_dimensions[args.level])
        downsample = float(slide.level_downsamples[args.level])
        polygons = _scaled_polygons(features, downsample)
        tissue = _preview_tissue(slide, preview_path)
        tissue_integral = _integral(tissue)
        preview_height, preview_width = tissue.shape
        rows = []
        for y in _grid_positions(level_height, args.patch_size, args.stride):
            for x in _grid_positions(level_width, args.patch_size, args.stride):
                left = max(0, min(preview_width - 1, math.floor(x * preview_width / level_width)))
                top = max(0, min(preview_height - 1, math.floor(y * preview_height / level_height)))
                right = max(left + 1, min(preview_width, math.ceil((x + args.patch_size) * preview_width / level_width)))
                bottom = max(top + 1, min(preview_height, math.ceil((y + args.patch_size) * preview_height / level_height)))
                tissue_fraction = _area_sum(tissue_integral, left, top, right, bottom) / (
                    (right - left) * (bottom - top)
                )
                mask = _binary_mask(polygons, x, y, args.patch_size)
                mask_array = np.asarray(mask, dtype=np.uint8)
                tumor_pixels = int(mask_array.sum())
                if tissue_fraction < args.minimum_tissue_fraction and tumor_pixels == 0:
                    continue
                patch_id = f"{slide_id}_L{args.level}_X{x:06d}_Y{y:06d}"
                image_path = slide_output / "images" / f"{patch_id}.jpg"
                mask_path = slide_output / "masks" / f"{patch_id}.png"
                if not _valid_image(image_path, args.patch_size):
                    rgba = slide.read_region(
                        x, y, args.patch_size, args.patch_size, level=args.level
                    )
                    _save_patch(Image.fromarray(rgba, mode="RGBA").convert("RGB"), image_path)
                if not _valid_image(mask_path, args.patch_size):
                    _save_mask(mask, mask_path)
                rows.append(
                    {
                        "slide_name_split": slide_id,
                        "slide_id": slide_id,
                        "patch_id": patch_id,
                        "filepath": str(image_path.resolve()),
                        "image_path": str(image_path.resolve()),
                        "mask_path": str(mask_path.resolve()),
                        "x": x,
                        "y": y,
                        "level": args.level,
                        "split": record["split"],
                        "has_tumor": int(tumor_pixels > 0),
                        "tumor_fraction": f"{tumor_pixels / (args.patch_size ** 2):.8f}",
                        "tissue_fraction": f"{tissue_fraction:.8f}",
                    }
                )
    if not rows:
        raise ValueError(f"Tissue screening produced no patches for {slide_id}")
    _atomic_csv(rows_path, rows)
    _atomic_json(
        marker_path,
        {
            "source_signature": source_signature,
            "patch_count": len(rows),
            "tumor_patch_count": sum(int(row["has_tumor"]) for row in rows),
        },
    )
    return rows


def _selection_key(seed: int, patch_id: str) -> str:
    return hashlib.sha256(f"{seed}|{patch_id}".encode()).hexdigest()


def _write_selections(rows: list[dict], output_root: Path, seed: int, negative_ratio: int) -> dict:
    by_split = defaultdict(list)
    for row in rows:
        by_split[row["split"]].append(row)
    selected = {"valid": by_split["valid"], "test": by_split["test"]}
    by_slide = defaultdict(list)
    for row in by_split["train"]:
        by_slide[row["slide_id"]].append(row)
    train = []
    for slide_id, slide_rows in sorted(by_slide.items()):
        positives = [row for row in slide_rows if int(row["has_tumor"])]
        negatives = [row for row in slide_rows if not int(row["has_tumor"])]
        negative_limit = min(
            len(negatives),
            max(256, negative_ratio * len(positives)) if positives else 256,
        )
        negatives.sort(key=lambda row: _selection_key(seed, row["patch_id"]))
        train.extend(positives)
        train.extend(negatives[:negative_limit])
    train.sort(key=lambda row: (row["slide_id"], int(row["y"]), int(row["x"])))
    selected["train"] = train
    selection_root = output_root / "decoder_selection"
    summary = {}
    for split in ("train", "valid", "test"):
        split_rows = selected[split]
        _atomic_csv(selection_root / f"{split}.csv", split_rows)
        summary[split] = {
            "patches": len(split_rows),
            "tumor_patches": sum(int(row["has_tumor"]) for row in split_rows),
            "slides": len({row["slide_id"] for row in split_rows}),
        }
    _atomic_json(selection_root / "summary.json", summary)
    return summary


def prepare(args) -> dict:
    ready, status = discover_ready(args)
    if not status["ready"]:
        raise RuntimeError(f"Partial data threshold has not been reached: {status}")
    args.output_root.mkdir(parents=True, exist_ok=True)
    cohort_path = args.output_root / "cohort.json"
    cohort_config = {
        "format_version": 1,
        "binary_label_map": status["label_map"],
        "level": args.level,
        "patch_size": args.patch_size,
        "stride": args.stride,
        "minimum_tissue_fraction": args.minimum_tissue_fraction,
        "selection_seed": args.selection_seed,
        "negative_ratio": args.negative_ratio,
    }
    if cohort_path.is_file():
        cohort = json.loads(cohort_path.read_text(encoding="utf-8"))
        if cohort.get("config") != cohort_config:
            raise ValueError("Existing frozen cohort uses different preparation settings")
        by_id = {record["slide_id"]: record for record in ready}
        missing = sorted(set(cohort["slide_ids"]).difference(by_id))
        if missing:
            raise RuntimeError(f"Frozen cohort sources are no longer complete: {missing}")
        records = [by_id[slide_id] for slide_id in cohort["slide_ids"]]
    else:
        records = ready
        cohort = {
            "config": cohort_config,
            "slide_ids": [record["slide_id"] for record in records],
            "split_counts": dict(Counter(record["split"] for record in records)),
        }
        _atomic_json(cohort_path, cohort)

    expected = _expected_sizes(args.download_manifest)
    rows = []
    if args.workers == 1:
        for index, record in enumerate(records, start=1):
            print(f"[{index}/{len(records)}] preparing {record['slide_id']}", flush=True)
            rows.extend(_process_slide(record, args, expected))
    else:
        print(
            f"Preparing {len(records)} slides with {args.workers} worker processes",
            flush=True,
        )
        with ProcessPoolExecutor(max_workers=args.workers) as executor:
            futures = {
                executor.submit(_process_slide, record, args, expected): record["slide_id"]
                for record in records
            }
            for index, future in enumerate(as_completed(futures), start=1):
                slide_id = futures[future]
                rows.extend(future.result())
                print(
                    f"[{index}/{len(records)}] prepared {slide_id}",
                    flush=True,
                )
    rows.sort(key=lambda row: (row["slide_id"], int(row["y"]), int(row["x"])))
    _atomic_csv(args.output_root / "all_patches.csv", rows)
    for split in ("train", "valid", "test"):
        _atomic_csv(
            args.output_root / f"all_patches_{split}.csv",
            [row for row in rows if row["split"] == split],
        )
    selection = _write_selections(rows, args.output_root, args.selection_seed, args.negative_ratio)
    summary = {
        "slides": len(records),
        "patches": len(rows),
        "tumor_patches": sum(int(row["has_tumor"]) for row in rows),
        "selection": selection,
    }
    _atomic_json(args.output_root / "preparation_summary.json", summary)
    return summary


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    for command in ("status", "prepare"):
        child = subparsers.add_parser(command)
        child.add_argument("--data-root", type=Path, required=True)
        child.add_argument("--download-manifest", type=Path, required=True)
        child.add_argument("--metadata-xlsx", type=Path, required=True)
        child.add_argument("--minimum-ready", type=int, default=32)
        child.add_argument("--minimum-train", type=int, default=16)
        child.add_argument("--minimum-valid", type=int, default=8)
        if command == "prepare":
            child.add_argument("--output-root", type=Path, required=True)
            child.add_argument("--level", type=int, default=1)
            child.add_argument("--patch-size", type=int, default=224)
            child.add_argument("--stride", type=int, default=224)
            child.add_argument("--minimum-tissue-fraction", type=float, default=0.2)
            child.add_argument("--selection-seed", type=int, default=42)
            child.add_argument("--negative-ratio", type=int, default=3)
            child.add_argument("--workers", type=int, default=1)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    for name in ("data_root", "download_manifest", "metadata_xlsx"):
        setattr(args, name, getattr(args, name).expanduser().resolve())
    if args.command == "status":
        _, status = discover_ready(args)
        print(json.dumps(status, ensure_ascii=False, sort_keys=True))
        raise SystemExit(0 if status["ready"] else 3)
    args.output_root = args.output_root.expanduser().resolve()
    if args.workers < 1:
        raise ValueError("workers must be positive")
    print(json.dumps(prepare(args), ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
